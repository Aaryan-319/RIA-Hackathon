import os
import time
import openpyxl

dirCheck = os.listdir("./")
localtime = str(time.asctime(time.localtime(time.time())))

#checking for database folder:
def folderCheck():
    if "DBMS" in dirCheck:
        print("STATUS: DATABASE FOLDER FOUND")
    else:
        os.mkdir("./DBMS")
        print("STATUS: DATABASE FOLDER CREATED")

#creating new dataset files:
def create(filename):
    try:
        if filename in dirCheck:
            print("ERR-001")
            print("ERR-NOTICE: THE DATASET ALREADY EXISTS.")

        elif filename not in dirCheck:
            wb = openpyxl.Workbook()
            wb.save(f'./DBMS/{filename}.xlsx')
            wb = openpyxl.load_workbook(f'./DBMS/{filename}.xlsx')
            wb.properties.title = filename
            wb.properties.category = filename
            wb.properties.keywords = filename
            wb.save(f'./DBMS/{filename}.xlsx')
            wb.close()

    except Exception as e:
        print("ERR-101")
        print("ERR-NOTICE: UNKNOWN ERROR OCCURRED")

#searching the database for datasets:
def search(dataset):
    if dataset in dirCheck:
        searchResult = dirCheck
        print(searchResult)
        print(f"RECORDS FOUND AT {dirCheck}/{searchResult}")
    else:
        print("ERR-010")
        print("ERR-NOTICE: NO SUCH FILE FOUND!")

if __name__ == '__main__':
    while True:
        print("DATABASE MANAGEMENT SYSTEM")
        folderCheck()
        query = input(str("> "))
        query = query.split(" ")
        if "create" == query[0] or "-c" == query[0]:
            if "new" == query[1] or "-n" == query[1]:
                if "database" == query[2] or "dataset" == query[2] or "-db" == query[2]:
                    if len(query) == 3:
                        filename = input(str("TOPIC: "))
                        create(filename)
                        print("STATUS: A NEW DATA ENTRY CREATED")

                    elif len(query) == 4:
                        print(f"STATUS: DATA ENTRY {filename} CREATED")

                    elif len(query) > 4:
                        print("ERR-100")
                        print("ERR-NOTICE: INVALID COMMAND")
# Checking for the DBMS folder
def folder_check():
    if "DBMS" in dirCheck:
        print("STATUS: DATABASE FOLDER FOUND")
    else:
        os.mkdir("./DBMS")
        print("STATUS: DATABASE FOLDER CREATED")

# Create a new dataset for users
def create_dataset(filename):
    try:
        if filename in dirCheck:
            print("ERR-001: Dataset already exists.")
        else:
            wb = openpyxl.Workbook()
            wb.save(f'./DBMS/{filename}.xlsx')
            wb = openpyxl.load_workbook(f'./DBMS/{filename}.xlsx')
            sheet = wb.active
            sheet.append(["Username", "Email", "Password", "Date of Birth", "Address", "City", "State", "Zip Code"])
            wb.save(f'./DBMS/{filename}.xlsx')
            wb.close()
            print(f"STATUS: {filename} dataset created with headers")
    except Exception as e:
        print("ERR-101: Unknown error occurred")

# Check for the existence of the user database
def check_user_database():
    if "users.xlsx" not in dirCheck:
        print("STATUS: User database not found. Creating database...")
        create_dataset("users")
    else:
        print("STATUS: User database found.")

# Add a new user to the database
def add_user(username, email, password, dob, address, city, state, zip_code):
    try:
        wb = openpyxl.load_workbook('./DBMS/users.xlsx')
        sheet = wb.active
        sheet.append([username, email, password, dob, address, city, state, zip_code])
        wb.save('./DBMS/users.xlsx')
        wb.close()
        print(f"STATUS: User {username} registered successfully")
    except Exception as e:
        print("ERR-102: Unable to add user")

# Check user credentials for login
def check_user_credentials(username, password):
    try:
        wb = openpyxl.load_workbook('./DBMS/users.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            if row[0] == username and row[2] == password:
                wb.close()
                return True
        wb.close()
        return False
    except Exception as e:
        print("ERR-103: Unable to check user credentials")

# Search for a dataset
def search_dataset(dataset):
    if dataset in dirCheck:
        print(f"RECORDS FOUND: {dataset}")
    else:
        print("ERR-010: No such file found!")

# Main program loop
if __name__ == '__main__':
    while True:
        print("DATABASE MANAGEMENT SYSTEM")
        folder_check()
        query = input(str("> "))
        query = query.split(" ")
        if "create" == query[0] or "-c" == query[0]:
            if "new" == query[1] or "-n" == query[1]:
                if "database" == query[2] or "dataset" == query[2] or "-db" == query[2]:
                    if len(query) == 3:
                        filename = input(str("TOPIC: "))
                        create_dataset(filename)
                        print("STATUS: A new data entry created")
                    elif len(query) == 4:
                        print(f"STATUS: Data entry {filename} created")
                    elif len(query) > 4:
                        print("ERR-100: Invalid command")
                elif "user" == query[2] or "-u" == query[2]:
                    username = input(str("USERNAME: "))
                    email = input(str("EMAIL: "))
                    password = input(str("PASSWORD: "))
                    dob = input(str("DATE OF BIRTH: "))
                    address = input(str("ADDRESS: "))
                    city = input(str("CITY: "))
                    state = input(str("STATE: "))
                    zip_code = input(str("ZIP CODE: "))
                    add_user(username, email, password, dob, address, city, state, zip_code)
                else:
                    print("ERR-100: Invalid command")
            else:
                print("ERR-100: Invalid command")
        elif "login" == query[0] or "-l" == query[0]:
            username = input(str("USERNAME: "))
            password = input(str("PASSWORD: "))
            if check_user_credentials(username, password):
                print("STATUS: User logged in successfully")
            else:
                print("ERR-104: Invalid username or password")
        elif "search" == query[0] or "-s" == query[0]:
            dataset = input(str("DATASET: "))
            search_dataset(dataset)
        else:
            print("ERR-100: Invalid command")
# Checking for the DBMS folder
def folderCheck():
    if "DBMS" in dirCheck:
        print("STATUS: DATABASE FOLDER FOUND")
    else:
        os.mkdir("./DBMS")
        print("STATUS: DATABASE FOLDER CREATED")

# Create a new dataset for users
def create(filename):
    try:
        if filename in dirCheck:
            print("ERR-001: Dataset already exists.")
        else:
            wb = openpyxl.Workbook()
            wb.save(f'./DBMS/{filename}.xlsx')
            wb = openpyxl.load_workbook(f'./DBMS/{filename}.xlsx')
            sheet = wb.active
            sheet.append(["Username", "Email", "Password", "Date of Birth", "Address", "City", "State", "Zip Code"])
            wb.save(f'./DBMS/{filename}.xlsx')
            wb.close()
            print(f"STATUS: {filename} dataset created with headers")
    except Exception as e:
        print("ERR-101: Unknown error occurred")

# Check for the existence of the user database
def check_user_database():
    if "users.xlsx" not in dirCheck:
        print("STATUS: User database not found. Creating database...")
        create("users")
    else:
        print("STATUS: User database found.")

# Add a new user to the database
def add_user(username, email, password, dob, address, city, state, zip_code):
    wb = openpyxl.load_workbook('./DBMS/users.xlsx')
    sheet = wb.active
    sheet.append([username, email, password, dob, address, city, state, zip_code])
    wb.save('./DBMS/users.xlsx')
    wb.close()
    print(f"STATUS: User {username} registered successfully")

# Check user credentials for login
def check_user_credentials(username, password):
    wb = openpyxl.load_workbook('./DBMS/users.xlsx')
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        if row[0] == username and row[2] == password:
            wb.close()
            return True
    wb.close()
    return False

# Search for a dataset
def search(dataset):
    if dataset in dirCheck:
        print(f"RECORDS FOUND: {dataset}")
    else:
        print("ERR-010: No such file found!")